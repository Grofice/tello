import openpyxl

NUMBER = 2 

f = open('savepath.txt')
f_data = f.readlines()
f_name = f_data[0]
f.close()


def createExcel():
    i = 1

    f2 = open('savepath.txt',mode='w')
    tmp = f_data[0].replace(f_data[0][10],str(int(f_data[0][10]) + 1)) 
    f2.write(tmp)
    f2.close()

    excel = openpyxl.Workbook()
    excel.create_sheet(f_name) 
    table = excel.active 

    data = ['number','width','high','area']

    for value in data:
        table.cell(1,i).value = value
        i = i + 1

    excel.save(f_name)


def writeExcel(data):
    i = 1
    excel = openpyxl.load_workbook(f_name)

    sheetnames = excel.get_sheet_names()
    table = excel.get_sheet_by_name(sheetnames[0])

    table = excel.active

    nrows = table.max_row
    ncolumns = table.max_column

    for value in data:
        table.cell(nrows+1,i).value = value
        i = i + 1
    
    excel.save(f_name)

if __name__ == "__main__":
    createExcel()
    data = ['1','2','3','4']
    writeExcel(data)
    writeExcel(data)
    # f = open('savepath.txt',mode='w+')
    # f_data = f.read()
    # f.close()
    # f_name = str(f_data)
    # for i in f_data[0]:
    #     # print(i)
    # print(f_data)
    
    # # print(type(f_data[0]))
