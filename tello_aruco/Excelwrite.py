import openpyxl

i = 1
# NUMBER = 40
NUMBER = 4
SAVE_PATH = './tello_fly/'
FILE_NAME = 'tello_fly'
FILE_NAME = 'MARKWandAREA'


excel = openpyxl.Workbook()
excel.create_sheet('Sheet1') 
table = excel.active 

# data = ['lr(x)','fb','ud(y)']
data = ['markw','area', 'catercorner']
j = 1
for value in data:
    table.cell(i,j).value = value
    j = j + 1

# excel.save('tello_fly'+str(NUMBER)+'.xlsx')


def createExcel():
    i = 1
    excel = openpyxl.Workbook()
    excel.create_sheet('Sheet1') 
    table = excel.active 

    data = ['number','width','high','area']

    for value in data:
        table.cell(1,i).value = value
        i = i + 1

    excel.save('excel_test'+str(NUMBER)+'.xlsx')


def writeExcel(data):
    j = 1
    global i
    i = i + 1
    for value in data:
        table.cell(i,j).value = value
        j = j + 1
    
    
    # excel.save('excel_test'+str(NUMBER)+'.xlsx')

if __name__ == "__main__":
    # createExcel()
    data = ['1','2','3','4']
    data = [1, 1, 1]
    writeExcel(data)
    writeExcel(data)
    writeExcel([1, 1, 1])
    excel.save('./tello_fly/tello_fly'+str(NUMBER)+'.xlsx')
    